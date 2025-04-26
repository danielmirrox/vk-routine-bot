from dotenv import load_dotenv
import os
import openpyxl
import datetime
import requests

load_dotenv()

book = openpyxl.load_workbook("schedule.xlsx")
sheet = book.active

days = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресение']
relative_days = ['вчера', 'сегодня', 'завтра']
today_idx = datetime.datetime.today().weekday()

def normalize_day(key):
    if key.lower() == 'сегодня':
        return days[today_idx]
    elif key.lower() == 'вчера':
        return days[today_idx - 1]
    elif key.lower() == 'завтра':
        return days[0] if today_idx == 6 else days[today_idx + 1]
    return key.lower()

def rasp(day_key, group, week_offset=0):
    key = normalize_day(day_key)
    group_col = find_group_column(group)
    if group_col is None:
        return []

    for day_idx, day_name in enumerate(days):
        if key == day_name:
            start_row = 4 + week_offset + day_idx * 14
            for row in range(start_row, start_row + 13, 2):
                yield sheet.cell(row=row, column=group_col).value

def find_group_column(group):
    for col in range(5, sheet.max_column + 1):
        value = sheet.cell(row=2, column=col).value
        if value and isinstance(value, str) and value.lower() == group.lower():
            return col
    return None

def prepods():
    prepod_list = []
    for col in range(8, 448, 10):
        for row in range(4, 87):
            cell = sheet.cell(row=row, column=col).value
            if isinstance(cell, str):
                names = cell.lower().replace('\n', '').replace('2', '').replace('1', '').replace(',', '').split('п/г')
                prepod_list.extend(name.strip() for name in names)
    return prepod_list

def rasp_prepod(day_key, prepod_name, week_offset=0):
    key = normalize_day(day_key)
    for col in range(8, 448, 5):
        lesson_num = 0
        for row in range(4 + week_offset, 87, 2):
            if (row - 4) % 14 == 0:
                lesson_num = 0
            lesson_num += 1
            cell = sheet.cell(row=row, column=col).value
            if isinstance(cell, str) and prepod_name in cell.lower():
                day_idx = (row - 4) // 14
                if days[day_idx] == key:
                    yield (lesson_num, day_idx, sheet.cell(row=row, column=col-2).value)

POGODA_TOKEN = os.getenv('POGODA_TOKEN')
WEATHER_URL = f'https://api.openweathermap.org/data/2.5/weather?q=Moscow&appid={POGODA_TOKEN}&lang=ru'
FORECAST_URL = 'https://api.openweathermap.org/data/2.5/forecast?q=Moscow'

def get_pogoda():
    response = requests.get(WEATHER_URL).json()
    yield f"https://openweathermap.org/img/wn/{response['weather'][0]['icon']}@2x.png"
    yield f"{response['weather'][0]['description'].capitalize()}, температура - {response['main']['temp']:.2f}°C"
    yield f"Давление - {response['main']['pressure']} мм рт. ст., влажность - {response['main']['humidity']}%"
    yield f"Ветер - {response['wind']['speed']} м/c"

def get_mnogo_pogoda():
    response = requests.get(FORECAST_URL, params={'units': 'metric', 'lang': 'ru', 'APPID': POGODA_TOKEN}).json()

    result = []
    current_day = ""

    for item in response['list']:
        forecast_time = datetime.datetime.fromtimestamp(item['dt'])
        day = forecast_time.strftime("%d.%m.%Y")
        time = forecast_time.strftime("%H:%M")

        if day != current_day:
            if current_day:
                result.append("\n")
            result.append(f"{day}:")
            current_day = day

        desc = item['weather'][0]['description'].capitalize()
        temp = item['main']['temp']
        result.append(f"{time} — {desc}, {temp:.2f}°C")

    return result

def get_weather_by_day(day: str):
    today = datetime.datetime.now().date()

    if day == 'сегодня':
        target_date = today
    elif day == 'завтра':
        target_date = today + datetime.timedelta(days=1)
    elif day == 'вчера':
        return ["К сожалению, данные за вчера недоступны."]
    else:
        return ["Неверный день для прогноза."]

    response = requests.get(FORECAST_URL, params={'units': 'metric', 'lang': 'ru', 'APPID': POGODA_TOKEN}).json()

    forecast = []
    for item in response['list']:
        forecast_time = datetime.datetime.fromtimestamp(item['dt']).date()
        if forecast_time == target_date:
            desc = item['weather'][0]['description']
            temp = item['main']['temp']
            forecast.append(f"{desc.capitalize()}, {temp:.2f}°C")

    if not forecast:
        return ["Нет данных для выбранной даты."]
    return forecast[:1]
